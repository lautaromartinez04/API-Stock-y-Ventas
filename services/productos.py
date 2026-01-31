# src/services/productos.py

from fastapi import HTTPException, status
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import load_workbook

from models.productos import Producto as ProductoModel
from schemas.producto import Producto, ProductoCreate

class ProductoService:
    model = ProductoModel

    def __init__(self, db: Session):
        self.db = db

    def get_all(self) -> list[Producto]:
        prods = self.db.query(ProductoModel).all()
        return [Producto.model_validate(p) for p in prods]

    def get(self, id: int) -> Producto | None:
        prod = self.db.query(ProductoModel).filter(ProductoModel.id == id).first()
        return prod and Producto.model_validate(prod)

    def create(self, payload: ProductoCreate) -> Producto:
        try:
            if payload.stock_bajo < 0:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "stock_bajo no puede ser negativo")
            if payload.precio_costo < 0:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "precio_costo no puede ser negativo")
            if payload.precio_unitario < payload.precio_costo:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "precio_unitario debe ser ≥ precio_costo")

            margen = (
                (payload.precio_unitario - payload.precio_costo) / payload.precio_costo * 100
                if payload.precio_costo > 0 else 0.0
            )

            prod = ProductoModel(
                nombre          = payload.nombre,
                codigo          = payload.codigo,
                descripcion     = payload.descripcion,
                stock_actual    = payload.stock_actual,
                stock_bajo      = payload.stock_bajo,
                precio_costo    = payload.precio_costo,
                margen          = margen,
                precio_unitario = payload.precio_unitario,
                categoria_id    = payload.categoria_id,
                activo          = payload.activo,
                image_url       = payload.image_url,
            )
            self.db.add(prod)
            self.db.commit()
            self.db.refresh(prod)
            return Producto.model_validate(prod)

        except HTTPException:
            self.db.rollback()
            raise
        except SQLAlchemyError:
            self.db.rollback()
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Error interno al crear el producto")

    def update(self, id: int, payload: ProductoCreate) -> Producto | None:
        prod = self.db.query(ProductoModel).filter(ProductoModel.id == id).first()
        if not prod:
            return None

        data = payload.model_dump(exclude_none=True)
        new_costo = data.get("precio_costo", prod.precio_costo)
        new_precio = data.get("precio_unitario", prod.precio_unitario)
        if new_costo < 0:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "precio_costo no puede ser negativo")
        if new_precio < new_costo:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "precio_unitario debe ser ≥ precio_costo")

        prod.precio_costo    = new_costo
        prod.precio_unitario = new_precio
        prod.margen = (
            (new_precio - new_costo) / new_costo * 100
            if new_costo > 0 else 0.0
        )

        if "stock_bajo" in data and data["stock_bajo"] < 0:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "stock_bajo no puede ser negativo")

        for fld in (
            "nombre", "codigo", "descripcion",
            "stock_actual", "stock_bajo",
            "categoria_id", "activo", "image_url"
        ):
            if fld in data:
                setattr(prod, fld, data[fld])

        try:
            self.db.commit()
            self.db.refresh(prod)
            return Producto.model_validate(prod)
        except SQLAlchemyError:
            self.db.rollback()
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Error interno al actualizar el producto")

    def delete(self, id: int) -> None:
        self.db.query(ProductoModel).filter(ProductoModel.id == id).delete()
        self.db.commit()

    def set_image(self, id: int, image_url: str) -> Producto:
        prod = self.db.query(ProductoModel).filter(ProductoModel.id == id).first()
        if not prod:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Producto no encontrado")
        prod.image_url = image_url
        self.db.commit()
        self.db.refresh(prod)
        return Producto.model_validate(prod)

    def bulk_update_prices_from_excel(self, file_bytes: bytes) -> tuple[int, list[str]]:
        try:
            wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except Exception:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "No se pudo leer el Excel")

        # 1) Leer la fila de encabezados
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            idx_id = headers.index("id")
            idx_price = headers.index("precio_unitario")
        except ValueError:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "El Excel debe tener encabezados 'id' y 'precio_unitario'"
            )

        updated = 0
        errors = []

        # 2) Iterar a partir de la fila 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            prod_id = row[idx_id]
            raw_precio = row[idx_price]

            if prod_id is None or raw_precio is None:
                errors.append(f"Fila {row_idx}: faltan datos en id o precio")
                continue

            # 3) Convertir precio a float
            try:
                nuevo_precio = float(raw_precio)
            except Exception:
                errors.append(f"Fila {row_idx}: precio inválido '{raw_precio}'")
                continue

            prod = self.db.query(ProductoModel).filter(ProductoModel.id == int(prod_id)).first()
            if not prod:
                errors.append(f"Fila {row_idx}: producto {prod_id} no existe")
                continue
            if nuevo_precio < prod.precio_costo:
                errors.append(f"Fila {row_idx}: precio_unitario menor al precio de costo")
                continue

            # 4) Actualizar
            prod.precio_unitario = nuevo_precio
            prod.margen = (
                (nuevo_precio - prod.precio_costo) / prod.precio_costo * 100
                if prod.precio_costo > 0 else 0.0
            )
            self.db.add(prod)
            updated += 1

        # 5) Commit
        try:
            self.db.commit()
        except SQLAlchemyError:
            self.db.rollback()
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR,
                                "Error interno al importar precios")

        return updated, errors

